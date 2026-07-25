# """
# Core engine for Excel Auto-Fill Bot.

# Responsibilities:
# - Read an Excel workbook (all sheets), auto-detect the header row for each sheet.
# - Parse a JSON or XML file into one or more flat "records" (key -> value).
# - Fuzzy-match each Excel header against the record's keys (headers do NOT need
#   to match exactly - "Customer Name" will match "customerName", "cust_name",
#   "customer.name", etc.)
# - Fill matched values into the first available empty row of each sheet that has
#   at least one matching header. Headers with no matching key are left blank.
# - Save back into the SAME excel file (overwrite), preserving all other sheets
#   and existing data untouched.

# No hardcoded headers anywhere - everything is auto-detected at runtime.
# """

# import json
# import os
# import re
# import xml.etree.ElementTree as ET
# from dataclasses import dataclass, field
# from typing import Any

# import openpyxl
# from rapidfuzz import fuzz

# # ---------------------------------------------------------------------------
# # Matching config
# # ---------------------------------------------------------------------------
# FUZZY_THRESHOLD = 72  # 0-100. Below this, a header is treated as "no match".
# BLANK_PLACEHOLDER = "N/A"  # written into any cell that has no matching/available data

# # Generic "qualifier" words that commonly prefix a field name without
# # changing its meaning - e.g. "Customer Name" and "Full Name" both really
# # just mean "Name". Stripping these lets a header like "Customer Name"
# # match a key like "personalDetails.fullName" even though only one of
# # their two words ("name") is literally shared. A distinguishing word that
# # is NOT a qualifier (e.g. "Product" in "Product Name") is left in place,
# # so it still correctly refuses to match a person's name field.
# QUALIFIER_WORDS = {
#     "full", "first", "last", "middle", "given", "customer", "client",
#     "user", "account", "holder", "contact", "primary", "legal", "official",
#     "person", "registered", "billing",
# }


# def normalize(text: str) -> str:
#     """Lowercase and strip everything except letters/digits for loose comparison."""
#     return re.sub(r"[^a-z0-9]", "", str(text).lower())


# # ---------------------------------------------------------------------------
# # Step 1: Parse JSON / XML into a list of flat records
# # ---------------------------------------------------------------------------

# def _flatten(obj: Any, parent_key: str = "", out: dict | None = None) -> dict:
#     """Flatten nested dict/list into dot-notation keys -> scalar values."""
#     if out is None:
#         out = {}

#     if isinstance(obj, dict):
#         for k, v in obj.items():
#             new_key = f"{parent_key}.{k}" if parent_key else str(k)
#             _flatten(v, new_key, out)
#     elif isinstance(obj, list):
#         if all(not isinstance(x, (dict, list)) for x in obj):
#             # list of scalars -> join into a single readable string
#             out[parent_key] = ", ".join(str(x) for x in obj)
#         else:
#             for i, item in enumerate(obj):
#                 _flatten(item, f"{parent_key}[{i}]", out)
#     else:
#         out[parent_key] = obj

#     return out


# def _extract_records(data: Any) -> list[dict]:
#     """
#     Given parsed JSON (dict/list), figure out whether it represents a single
#     record or multiple records, and return a list of flat dicts.
#     """
#     # Top-level list -> each item is a record
#     if isinstance(data, list):
#         return [_flatten(item) for item in data]

#     if isinstance(data, dict):
#         # Common wrapper pattern: {"records": [...]}  or {"items": [...]}
#         list_valued_keys = [
#             k for k, v in data.items()
#             if isinstance(v, list) and v and all(isinstance(x, dict) for x in v)
#         ]
#         # Only treat as "records wrapper" if that's basically the whole payload
#         if len(list_valued_keys) == 1 and len(data) == 1:
#             return [_flatten(item) for item in data[list_valued_keys[0]]]

#         # Otherwise it's a single record
#         return [_flatten(data)]

#     return [{"value": data}]


# def _local_name(tag: str) -> str:
#     """Strip a '{namespace}' prefix off an ElementTree tag, leaving just the local name."""
#     return tag.split("}", 1)[-1] if tag.startswith("{") else tag


# def _xml_elem_to_obj(elem: ET.Element):
#     """Recursively convert an XML element into dict/list/scalar (like JSON)."""
#     children = list(elem)
#     if not children:
#         text = (elem.text or "").strip()
#         node: Any = text if text != "" else None
#         if elem.attrib:
#             merged = {f"@{_local_name(k)}": v for k, v in elem.attrib.items()}
#             if node is not None:
#                 merged["#text"] = node
#             return merged
#         return node

#     result: dict[str, Any] = {}
#     for child in children:
#         child_val = _xml_elem_to_obj(child)
#         tag = _local_name(child.tag)
#         if tag in result:
#             if not isinstance(result[tag], list):
#                 result[tag] = [result[tag]]
#             result[tag].append(child_val)
#         else:
#             result[tag] = child_val

#     if elem.attrib:
#         for k, v in elem.attrib.items():
#             result[f"@{_local_name(k)}"] = v

#     return result


# def _strip_all_namespace_prefixes(xml_text: str) -> str:
#     """
#     Aggressively remove all namespace prefixes and declarations so that
#     expat can parse the XML even when prefixes are unbound.
#     Also normalises common XML malformations (escaped quotes in xmlns,
#     spaces inside tag names).
#     """
#     # 0. Normalise backslash-escaped quotes in xmlns declarations so the
#     #    subsequent regexes can match them. Handles \" and \\" patterns.
#     xml_text = re.sub(
#         r'(xmlns(?::[A-Za-z_][\w.\-]*)?\s*=\s*)\\*"',
#         r'\1"',
#         xml_text,
#     )
#     # Also handle the case where xmlns= already had its prefix stripped by
#     # a prior pass – i.e. just ns4=\\"...\\" (without leading xmlns:).
#     # We catch stray prefix:attr patterns that are NOT XML-safe.
#     xml_text = re.sub(
#         r'(?<=\s)(ns\d+)\s*=\s*\\+".*?\\+"',
#         '',
#         xml_text,
#     )

#     # 1. Remove all xmlns:prefix and default xmlns declarations
#     xml_text = re.sub(r'xmlns:[A-Za-z_][\w.\-]*\s*=\s*"[^"]*"', '', xml_text)
#     xml_text = re.sub(r"xmlns:[A-Za-z_][\w.\-]*\s*=\s*'[^']*'", '', xml_text)
#     xml_text = re.sub(r'xmlns\s*=\s*"[^"]*"', '', xml_text)
#     xml_text = re.sub(r"xmlns\s*=\s*'[^']*'", '', xml_text)

#     # 2. Strip prefixes from opening tags: <prefix:tag -> <tag
#     xml_text = re.sub(r'<[A-Za-z_][\w.\-]*:', '<', xml_text)
#     # 3. Strip prefixes from closing tags: </prefix:tag -> </tag
#     xml_text = re.sub(r'</[A-Za-z_][\w.\-]*:', '</', xml_text)
#     # 4. Strip prefixes from attribute names: \s+prefix:attr= -> attr=
#     xml_text = re.sub(r'\s+[A-Za-z_][\w.\-]*:', ' ', xml_text)

#     # 5. Extra safety: remove any remaining colon that appears right after < or </
#     xml_text = re.sub(r'(?<=<)\w+:', '', xml_text)
#     xml_text = re.sub(r'(?<=</)\w+:', '', xml_text)

#     # 6. Fix spaces inside tag names – e.g.
#     #    <returned NewMember> -> <returnedNewMember>
#     #    <deceased Indicator> -> <deceasedIndicator>
#     #    < postalCode> -> <postalCode>
#     #    (prefix already stripped by step 2, but the space after < remains)
#     xml_text = re.sub(r'<(\s+)(\w)', r'<\2', xml_text)
#     xml_text = re.sub(r'</(\s+)(\w)', r'</\2', xml_text)
#     xml_text = re.sub(r'<(\w+)\s+(\w[\w.-]*)([>\s])', r'<\1\2\3', xml_text)
#     xml_text = re.sub(r'</(\w+)\s+(\w[\w.-]*)>', r'</\1\2>', xml_text)

#     return xml_text


# def _load_xml_text(path: str) -> str:
#     """Read an XML file as text, tolerating a UTF-8 BOM and stray leading bytes."""
#     with open(path, "r", encoding="utf-8-sig") as f:
#         text = f.read()
#     # Defensive: drop anything before the first '<' (stray whitespace/junk
#     # that can confuse the parser into misreporting the real error location).
#     first_lt = text.find("<")
#     if first_lt > 0:
#         text = text[first_lt:]
#     return text


# def _extract_xml_with_regex(xml_text: str) -> list[dict]:
#     """
#     Last-resort extraction from broken XML that can't be parsed by ElementTree.
#     Uses regex to find <tag>value</tag> patterns and returns a single flat record.
#     """
#     # Strip namespace prefixes first (same as _strip_all_namespace_prefixes)
#     text = _strip_all_namespace_prefixes(xml_text)

#     # Find all <tag>scalar_value</tag> patterns (non-nested, self-closing excluded)
#     pattern = re.compile(r'<(\w+)>([^<]+)</\1>')
#     matches = pattern.findall(text)

#     record: dict[str, list[str]] = {}
#     for tag, value in matches:
#         v = value.strip()
#         if v:
#             record.setdefault(tag, []).append(v)

#     # Convert multi-valued fields to comma-separated strings
#     flat = {tag: ", ".join(vals) if len(vals) > 1 else vals[0]
#             for tag, vals in record.items()}

#     return [flat] if flat else []


# def parse_data_file(path: str) -> list[dict]:
#     """Load a .json or .xml file and return a list of flat records."""
#     lower = path.lower()

#     if lower.endswith(".json"):
#         with open(path, "r", encoding="utf-8-sig") as f:
#             raw = json.load(f)
#         return _extract_records(raw)

#     if lower.endswith(".xml"):
#         xml_text = _load_xml_text(path)

#         try:
#             root = ET.fromstring(xml_text)
#         except ET.ParseError as first_err:
#             # Retry, stripping namespace prefixes outright.
#             try:
#                 stripped = _strip_all_namespace_prefixes(xml_text)
#                 root = ET.fromstring(stripped)
#             except ET.ParseError:
#                 # Final fallback: regex-based extraction for fundamentally
#                 # broken XML (mismatched tags, spacing inside tag names, etc.)
#                 records = _extract_xml_with_regex(xml_text)
#                 if records:
#                     return records
#                 raise ValueError(
#                     f"Could not parse XML file '{path}': {first_err}. "
#                     "The file may be malformed or truncated."
#                 ) from first_err

#         obj = _xml_elem_to_obj(root)

#         # If the root's children are all the same repeated tag, treat them as records
#         # e.g. <records><record>...</record><record>...</record></records>
#         if isinstance(obj, dict):
#             list_valued_keys = [
#                 k for k, v in obj.items()
#                 if isinstance(v, list) and v and all(isinstance(x, dict) for x in v)
#             ]
#             if len(list_valued_keys) == 1 and len(obj) == 1:
#                 return [_flatten(item) for item in obj[list_valued_keys[0]]]

#         return _extract_records(obj)

#     raise ValueError(f"Unsupported data file type: {path} (expected .json or .xml)")


# # ---------------------------------------------------------------------------
# # Step 2: Auto-detect header row + fill logic
# # ---------------------------------------------------------------------------

# @dataclass
# class SheetPlan:
#     sheet_name: str
#     header_row: int
#     headers: dict[int, str] = field(default_factory=dict)  # col_idx -> header text
#     matches: dict[int, str] = field(default_factory=dict)  # col_idx -> matched record key
#     unmatched_headers: list[str] = field(default_factory=list)


# def _detect_header_row(ws, max_scan: int = 5) -> int:
#     """Return the 1-indexed row number most likely to be the header row."""
#     best_row, best_score = 1, -1
#     for r in range(1, min(max_scan, ws.max_row) + 1):
#         cells = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
#         non_empty = [c for c in cells if c not in (None, "")]
#         if not non_empty:
#             continue
#         text_like = sum(1 for c in non_empty if isinstance(c, str))
#         score = text_like / len(non_empty) * len(non_empty)
#         if score > best_score:
#             best_score, best_row = score, r
#     return best_row


# def _first_empty_row(ws, header_row: int) -> int:
#     """First row (after header_row) where every cell is blank."""
#     r = header_row + 1
#     while r <= ws.max_row:
#         row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
#         if all(v in (None, "") for v in row_vals):
#             return r
#         r += 1
#     return max(r, header_row + 1)


# def _tokenize(text: str) -> list[str]:
#     """Split a header/key into lowercase word tokens, handling camelCase,
#     snake_case, kebab-case, and spaces. 'panNumber' -> ['pan', 'number']."""
#     s = str(text)
#     s = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", s)  # camelCase boundary
#     s = re.sub(r"[^A-Za-z0-9]+", " ", s)
#     return [t.lower() for t in s.split() if t]


# def _best_match(header: str, record: dict[str, Any]) -> str | None:
#     """
#     Find the record key that best matches an Excel header. Returns None if
#     nothing clears the confidence bar.

#     Strategy (in order of trust):
#     1. Exact match on the normalized full string.
#     2. Acronym match - e.g. header "dob" matches key "date_of_birth" because
#        its word-initials spell "dob".
#     3. Token overlap - header and key are split into words (camelCase/
#        snake_case/space aware) and matched by how much of the *smaller*
#        token set is covered by shared words. This is what lets "Email
#        Address" match "email" and "Customer Name" match "customerName",
#        while correctly rejecting things like "panNumber" vs "mobileNumber"
#        (they only share the word "number", not enough coverage).
#     4. Plain character-similarity fallback - only for single-word vs
#        single-word comparisons, to tolerate small typos/spelling variants.
#     """
#     norm_header = normalize(header)
#     if not norm_header:
#         return None

#     header_tokens = _tokenize(header)
#     best_key, best_score = None, 0

#     for key in record:
#         leaf = key.split(".")[-1]
#         leaf = re.sub(r"\[\d+\]$", "", leaf)
#         norm_leaf = normalize(leaf)
#         norm_full = normalize(key)

#         if norm_header == norm_leaf or norm_header == norm_full:
#             return key  # exact match, short circuit

#         leaf_tokens = _tokenize(leaf)

#         # Acronym match: "dob" <-> ["date", "of", "birth"]
#         if len(leaf_tokens) >= 2 and norm_header == "".join(t[0] for t in leaf_tokens):
#             return key
#         if len(header_tokens) >= 2 and norm_leaf == "".join(t[0] for t in header_tokens):
#             return key

#         # Qualifier-stripped match: drop generic qualifier words ("full",
#         # "customer", "contact", ...) from BOTH sides and check what's left
#         # is exactly equal. This is what lets "Customer Name" match
#         # "personalDetails.fullName" (both reduce to just {"name"}), while
#         # still correctly rejecting "Product Name" (reduces to
#         # {"product", "name"} - "product" isn't a qualifier, so it stays
#         # and breaks the equality, leaving it unmatched as it should be).
#         h_core = frozenset(t for t in header_tokens if t not in QUALIFIER_WORDS)
#         l_core = frozenset(t for t in leaf_tokens if t not in QUALIFIER_WORDS)
#         if h_core and h_core == l_core:
#             return key

#         # Token overlap: how much of the smaller word-set is shared
#         hset, lset = set(header_tokens), set(leaf_tokens)
#         overlap = hset & lset
#         token_score = 0
#         if overlap:
#             coverage = len(overlap) / min(len(hset), len(lset))
#             if coverage >= 0.6:
#                 token_score = coverage * 100

#         # Character-similarity fallback: only trusted when BOTH sides are a
#         # single word each, so we don't get fooled by shared word-fragments
#         # across multi-word keys (e.g. "...Number" suffix collisions).
#         str_score = 0
#         if len(header_tokens) == 1 and len(leaf_tokens) == 1 and min(len(norm_header), len(norm_leaf)) >= 4:
#             str_score = max(
#                 fuzz.ratio(norm_header, norm_leaf),
#                 fuzz.partial_ratio(norm_header, norm_leaf) * 0.9,
#             )

#         score = max(token_score, str_score)
#         if score > best_score:
#             best_score, best_key = score, key

#     if best_score >= FUZZY_THRESHOLD:
#         return best_key
#     return None


# # Header tokens that usually mean "this column is a unique identifier" -
# # checked in priority order (most trustworthy first) when deciding which
# # column(s) to use for spotting a duplicate record.
# _IDENTIFIER_TOKENS = [
#     "uid", "aadhar", "aadhaar", "pan", "id",
#     "email", "account", "number", "code", "mobile", "phone",
# ]


# def _pick_key_columns(headers: dict[int, str], matches: dict[int, str]) -> list[int]:
#     """
#     Decide which matched column(s) identify "the same record" for duplicate
#     detection. Prefers a single obviously-unique column (uid, aadhaar, pan,
#     id, email, account/number/code, mobile/phone - in that trust order).
#     Falls back to using ALL matched columns together (composite key) when no
#     such column is present.
#     """
#     for token in _IDENTIFIER_TOKENS:
#         for col in matches:
#             if token in _tokenize(headers.get(col, "")):
#                 return [col]
#     return list(matches.keys())


# def _norm_cell(value: Any) -> str:
#     """Normalize a cell value for duplicate comparison."""
#     if value in (None, ""):
#         return ""
#     return str(value).strip().lower()


# def _existing_row_index(ws, header_row: int, key_cols: list[int]) -> dict[tuple, int]:
#     """
#     Scan already-filled rows below the header and build a lookup of
#     key-values -> row number, so a new record that matches an existing row's
#     key can REPLACE it instead of being appended as a fresh duplicate row.
#     """
#     index: dict[tuple, int] = {}
#     if not key_cols:
#         return index

#     r = header_row + 1
#     while r <= ws.max_row:
#         row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
#         if any(v not in (None, "") for v in row_vals):
#             key = tuple(_norm_cell(ws.cell(row=r, column=c).value) for c in key_cols)
#             if any(key):  # ignore all-blank/N-A keys, nothing reliable to match on
#                 index[key] = r
#         r += 1
#     return index


# def build_plan(ws, record: dict[str, Any]) -> SheetPlan:
#     """Detect headers for one sheet and map them against one record."""
#     header_row = _detect_header_row(ws)
#     plan = SheetPlan(sheet_name=ws.title, header_row=header_row)

#     for c in range(1, ws.max_column + 1):
#         val = ws.cell(row=header_row, column=c).value
#         if val in (None, ""):
#             continue
#         header_text = str(val)
#         plan.headers[c] = header_text
#         match_key = _best_match(header_text, record)
#         if match_key is not None:
#             plan.matches[c] = match_key
#         else:
#             plan.unmatched_headers.append(header_text)

#     return plan


# def fill_workbook(
#     excel_path: str,
#     data_path: str,
#     save_path: str | None = None,
#     skip_unrelated_sheets: bool = True,
# ) -> dict:
#     """
#     Main entry point.

#     - A response only fills the sheet(s) it actually relates to. If a
#       workbook has many sheets (e.g. an e-commerce file with "Customers",
#       "Orders", "Products"...), a customer-details response fills only the
#       matching sheet - other sheets stay untouched. Feed a different
#       response later (Orders API, Products API, ...) and IT fills its own
#       matching sheet, same workbook. Set `skip_unrelated_sheets=False` to
#       instead force every sheet to get a row (unmatched cells become "N/A").
#     - Any header with no corresponding value in the response is written as
#       "N/A".
#     - Duplicate records REPLACE the existing row instead of adding a new
#       one. "Duplicate" is decided by a unique-looking column if the sheet
#       has one (uid / aadhaar / pan / id / email / account / number / code /
#       mobile - whichever appears first), otherwise by all matched columns
#       together. A genuinely new record still gets a brand new row.

#     Returns a summary dict describing what was matched/filled/replaced per
#     sheet, for display in the UI / logs. Saves the workbook to `save_path`
#     (defaults to overwriting `excel_path`).
#     """
#     save_path = save_path or excel_path
#     records = parse_data_file(data_path)
#     if not records:
#         raise ValueError("No records found in the JSON/XML file.")

#     wb = openpyxl.load_workbook(excel_path)
#     summary = {"records_processed": len(records), "sheets": []}

#     for sheet_name in wb.sheetnames:
#         ws = wb[sheet_name]
#         if ws.max_row == 0 or ws.max_column == 0:
#             continue

#         sheet_summary = {
#             "sheet": sheet_name,
#             "matched_headers": {},
#             "unmatched_headers": [],
#             "rows_filled": 0,
#             "rows_added": 0,
#             "rows_replaced": 0,
#             "skipped": False,
#         }

#         # Use the first record just to detect the header row / column layout
#         # and to judge whether this sheet is even relevant to this response.
#         first_plan = build_plan(ws, records[0])
#         if not first_plan.headers:
#             # Truly empty header row - nothing to key off, leave it alone.
#             sheet_summary["skipped"] = True
#             summary["sheets"].append(sheet_summary)
#             continue

#         sheet_is_relevant = any(build_plan(ws, r).matches for r in records)
#         if skip_unrelated_sheets and not sheet_is_relevant:
#             sheet_summary["skipped"] = True
#             sheet_summary["unmatched_headers"] = first_plan.unmatched_headers
#             summary["sheets"].append(sheet_summary)
#             continue

#         header_row = first_plan.header_row
#         key_cols = _pick_key_columns(first_plan.headers, first_plan.matches)
#         key_to_row = _existing_row_index(ws, header_row, key_cols)
#         next_row = _first_empty_row(ws, header_row)
#         ever_matched: dict[str, str] = {}  # header_text -> key (union across all records)

#         for record in records:
#             plan = build_plan(ws, record)

#             row_values: dict[int, Any] = {}
#             for col_idx, header_text in plan.headers.items():
#                 if col_idx in plan.matches:
#                     key = plan.matches[col_idx]
#                     value = record.get(key)
#                     if value is None or (isinstance(value, str) and value.strip() == ""):
#                         value = BLANK_PLACEHOLDER
#                     row_values[col_idx] = value
#                     ever_matched[header_text] = key
#                 else:
#                     row_values[col_idx] = BLANK_PLACEHOLDER

#             record_key = tuple(_norm_cell(row_values.get(c)) for c in key_cols) if key_cols else ()
#             target_row = key_to_row.get(record_key) if record_key and any(record_key) else None

#             if target_row is not None:
#                 row = target_row
#                 sheet_summary["rows_replaced"] += 1
#             else:
#                 row = next_row
#                 next_row += 1
#                 sheet_summary["rows_added"] += 1

#             for col_idx, value in row_values.items():
#                 ws.cell(row=row, column=col_idx, value=value)

#             if record_key and any(record_key):
#                 key_to_row[record_key] = row

#             sheet_summary["rows_filled"] += 1

#         sheet_summary["matched_headers"] = ever_matched
#         sheet_summary["unmatched_headers"] = [
#             h for h in first_plan.headers.values() if h not in ever_matched
#         ]
#         summary["sheets"].append(sheet_summary)

#     try:
#         wb.save(save_path)
#         summary["saved_to"] = save_path
#     except PermissionError:
#         # Usually means the Excel file is currently open in Excel (or is
#         # read-only). Don't lose the work - save alongside it instead.
#         base, ext = os.path.splitext(save_path)
#         fallback_path = f"{base}_filled{ext}"
#         n = 2
#         while os.path.exists(fallback_path):
#             fallback_path = f"{base}_filled{n}{ext}"
#             n += 1
#         try:
#             wb.save(fallback_path)
#         except PermissionError as e:
#             raise PermissionError(
#                 f"Could not save to '{save_path}' or '{fallback_path}'. "
#                 f"Close the Excel file if it's open, or check that the "
#                 f"Downloads folder allows write access, then try again."
#             ) from e
#         summary["saved_to"] = fallback_path
#         summary["save_fallback"] = True

#     return summary


































"""
Core engine for Excel Auto-Fill Bot.

Responsibilities:
- Read an Excel workbook (all sheets), auto-detect the header row for each sheet.
- Parse a JSON or XML file into one or more flat "records" (key -> value).
- Fuzzy-match each Excel header against the record's keys (headers do NOT need
  to match exactly - "Customer Name" will match "customerName", "cust_name",
  "customer.name", etc.)
- Fill matched values into the first available empty row of each sheet that has
  at least one matching header. Headers with no matching key are left blank.
- Save back into the SAME excel file (overwrite), preserving all other sheets
  and existing data untouched.

No hardcoded headers anywhere - everything is auto-detected at runtime.
"""

import json
import os
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from rapidfuzz import fuzz

# ---------------------------------------------------------------------------
# Matching config
# ---------------------------------------------------------------------------
FUZZY_THRESHOLD = 72  # 0-100. Below this, a header is treated as "no match".
BLANK_PLACEHOLDER = "N/A"  # written into any cell that has no matching/available data

# Generic "qualifier" words that commonly prefix a field name without
# changing its meaning - e.g. "Customer Name" and "Full Name" both really
# just mean "Name". Stripping these lets a header like "Customer Name"
# match a key like "personalDetails.fullName" even though only one of
# their two words ("name") is literally shared. A distinguishing word that
# is NOT a qualifier (e.g. "Product" in "Product Name") is left in place,
# so it still correctly refuses to match a person's name field.
#
# IMPORTANT: "first"/"last"/"middle"/"given" are intentionally NOT included
# here. They used to be, which meant "First Name", "Last Name", and
# "Full Name" all reduced to the exact same {"name"} core and matched
# whichever name-like key came first in the data - silently duplicating
# one person's first name into the Full Name AND Last Name columns. These
# words denote a specific, non-interchangeable part of a name, so they're
# handled separately (see _name_role / _augment_name_fields below) instead
# of being stripped as generic filler.
QUALIFIER_WORDS = {
    "full", "customer", "client",
    "user", "account", "holder", "contact", "primary", "legal", "official",
    "person", "registered", "billing",
}


def normalize(text: str) -> str:
    """Lowercase and strip everything except letters/digits for loose comparison."""
    return re.sub(r"[^a-z0-9]", "", str(text).lower())


# ---------------------------------------------------------------------------
# Step 1: Parse JSON / XML into a list of flat records
# ---------------------------------------------------------------------------

def _flatten(obj: Any, parent_key: str = "", out: dict | None = None) -> dict:
    """Flatten nested dict/list into dot-notation keys -> scalar values."""
    if out is None:
        out = {}

    if isinstance(obj, dict):
        for k, v in obj.items():
            new_key = f"{parent_key}.{k}" if parent_key else str(k)
            _flatten(v, new_key, out)
    elif isinstance(obj, list):
        if all(not isinstance(x, (dict, list)) for x in obj):
            # list of scalars -> join into a single readable string
            out[parent_key] = ", ".join(str(x) for x in obj)
        else:
            for i, item in enumerate(obj):
                _flatten(item, f"{parent_key}[{i}]", out)
    else:
        out[parent_key] = obj

    return out


def _extract_records(data: Any) -> list[dict]:
    """
    Given parsed JSON (dict/list), figure out whether it represents a single
    record or multiple records, and return a list of flat dicts.
    """
    # Top-level list -> each item is a record
    if isinstance(data, list):
        return [_flatten(item) for item in data]

    if isinstance(data, dict):
        # Common wrapper pattern: {"records": [...]}  or {"items": [...]}
        list_valued_keys = [
            k for k, v in data.items()
            if isinstance(v, list) and v and all(isinstance(x, dict) for x in v)
        ]
        # Only treat as "records wrapper" if that's basically the whole payload
        if len(list_valued_keys) == 1 and len(data) == 1:
            return [_flatten(item) for item in data[list_valued_keys[0]]]

        # Otherwise it's a single record
        return [_flatten(data)]

    return [{"value": data}]


def _local_name(tag: str) -> str:
    """Strip a '{namespace}' prefix off an ElementTree tag, leaving just the local name."""
    return tag.split("}", 1)[-1] if tag.startswith("{") else tag


def _xml_elem_to_obj(elem: ET.Element):
    """Recursively convert an XML element into dict/list/scalar (like JSON)."""
    children = list(elem)
    if not children:
        text = (elem.text or "").strip()
        node: Any = text if text != "" else None
        if elem.attrib:
            merged = {f"@{_local_name(k)}": v for k, v in elem.attrib.items()}
            if node is not None:
                merged["#text"] = node
            return merged
        return node

    result: dict[str, Any] = {}
    for child in children:
        child_val = _xml_elem_to_obj(child)
        tag = _local_name(child.tag)
        if tag in result:
            if not isinstance(result[tag], list):
                result[tag] = [result[tag]]
            result[tag].append(child_val)
        else:
            result[tag] = child_val

    if elem.attrib:
        for k, v in elem.attrib.items():
            result[f"@{_local_name(k)}"] = v

    return result


def _strip_all_namespace_prefixes(xml_text: str) -> str:
    """
    Aggressively remove all namespace prefixes and declarations so that
    expat can parse the XML even when prefixes are unbound.
    Also normalises common XML malformations (escaped quotes in xmlns,
    spaces inside tag names).
    """
    # 0. Normalise backslash-escaped quotes in xmlns declarations so the
    #    subsequent regexes can match them. Handles \" and \\" patterns.
    xml_text = re.sub(
        r'(xmlns(?::[A-Za-z_][\w.\-]*)?\s*=\s*)\\*"',
        r'\1"',
        xml_text,
    )
    # Also handle the case where xmlns= already had its prefix stripped by
    # a prior pass – i.e. just ns4=\\"...\\" (without leading xmlns:).
    # We catch stray prefix:attr patterns that are NOT XML-safe.
    xml_text = re.sub(
        r'(?<=\s)(ns\d+)\s*=\s*\\+".*?\\+"',
        '',
        xml_text,
    )

    # 1. Remove all xmlns:prefix and default xmlns declarations
    xml_text = re.sub(r'xmlns:[A-Za-z_][\w.\-]*\s*=\s*"[^"]*"', '', xml_text)
    xml_text = re.sub(r"xmlns:[A-Za-z_][\w.\-]*\s*=\s*'[^']*'", '', xml_text)
    xml_text = re.sub(r'xmlns\s*=\s*"[^"]*"', '', xml_text)
    xml_text = re.sub(r"xmlns\s*=\s*'[^']*'", '', xml_text)

    # 2. Strip prefixes from opening tags: <prefix:tag -> <tag
    xml_text = re.sub(r'<[A-Za-z_][\w.\-]*:', '<', xml_text)
    # 3. Strip prefixes from closing tags: </prefix:tag -> </tag
    xml_text = re.sub(r'</[A-Za-z_][\w.\-]*:', '</', xml_text)
    # 4. Strip prefixes from attribute names: \s+prefix:attr= -> attr=
    xml_text = re.sub(r'\s+[A-Za-z_][\w.\-]*:', ' ', xml_text)

    # 5. Extra safety: remove any remaining colon that appears right after < or </
    xml_text = re.sub(r'(?<=<)\w+:', '', xml_text)
    xml_text = re.sub(r'(?<=</)\w+:', '', xml_text)

    # 6. Fix spaces inside tag names – e.g.
    #    <returned NewMember> -> <returnedNewMember>
    #    <deceased Indicator> -> <deceasedIndicator>
    #    < postalCode> -> <postalCode>
    #    (prefix already stripped by step 2, but the space after < remains)
    xml_text = re.sub(r'<(\s+)(\w)', r'<\2', xml_text)
    xml_text = re.sub(r'</(\s+)(\w)', r'</\2', xml_text)
    xml_text = re.sub(r'<(\w+)\s+(\w[\w.-]*)([>\s])', r'<\1\2\3', xml_text)
    xml_text = re.sub(r'</(\w+)\s+(\w[\w.-]*)>', r'</\1\2>', xml_text)

    return xml_text


def _load_xml_text(path: str) -> str:
    """Read an XML file as text, tolerating a UTF-8 BOM and stray leading bytes."""
    with open(path, "r", encoding="utf-8-sig") as f:
        text = f.read()
    # Defensive: drop anything before the first '<' (stray whitespace/junk
    # that can confuse the parser into misreporting the real error location).
    first_lt = text.find("<")
    if first_lt > 0:
        text = text[first_lt:]
    return text


def _extract_xml_with_regex(xml_text: str) -> list[dict]:
    """
    Last-resort extraction from broken XML that can't be parsed by ElementTree.
    Uses regex to find <tag>value</tag> patterns and returns a single flat record.
    """
    # Strip namespace prefixes first (same as _strip_all_namespace_prefixes)
    text = _strip_all_namespace_prefixes(xml_text)

    # Find all <tag>scalar_value</tag> patterns (non-nested, self-closing excluded)
    pattern = re.compile(r'<(\w+)>([^<]+)</\1>')
    matches = pattern.findall(text)

    record: dict[str, list[str]] = {}
    for tag, value in matches:
        v = value.strip()
        if v:
            record.setdefault(tag, []).append(v)

    # Convert multi-valued fields to comma-separated strings
    flat = {tag: ", ".join(vals) if len(vals) > 1 else vals[0]
            for tag, vals in record.items()}

    return [flat] if flat else []


def parse_data_file(path: str) -> list[dict]:
    """Load a .json or .xml file and return a list of flat records."""
    lower = path.lower()

    if lower.endswith(".json"):
        with open(path, "r", encoding="utf-8-sig") as f:
            raw = json.load(f)
        return _extract_records(raw)

    if lower.endswith(".xml"):
        xml_text = _load_xml_text(path)

        try:
            root = ET.fromstring(xml_text)
        except ET.ParseError as first_err:
            # Retry, stripping namespace prefixes outright.
            try:
                stripped = _strip_all_namespace_prefixes(xml_text)
                root = ET.fromstring(stripped)
            except ET.ParseError:
                # Final fallback: regex-based extraction for fundamentally
                # broken XML (mismatched tags, spacing inside tag names, etc.)
                records = _extract_xml_with_regex(xml_text)
                if records:
                    return records
                raise ValueError(
                    f"Could not parse XML file '{path}': {first_err}. "
                    "The file may be malformed or truncated."
                ) from first_err

        obj = _xml_elem_to_obj(root)

        # If the root's children are all the same repeated tag, treat them as records
        # e.g. <records><record>...</record><record>...</record></records>
        if isinstance(obj, dict):
            list_valued_keys = [
                k for k, v in obj.items()
                if isinstance(v, list) and v and all(isinstance(x, dict) for x in v)
            ]
            if len(list_valued_keys) == 1 and len(obj) == 1:
                return [_flatten(item) for item in obj[list_valued_keys[0]]]

        return _extract_records(obj)

    raise ValueError(f"Unsupported data file type: {path} (expected .json or .xml)")


# ---------------------------------------------------------------------------
# Step 2: Auto-detect header row + fill logic
# ---------------------------------------------------------------------------

@dataclass
class SheetPlan:
    sheet_name: str
    header_row: int
    headers: dict[int, str] = field(default_factory=dict)  # col_idx -> header text
    matches: dict[int, str] = field(default_factory=dict)  # col_idx -> matched record key
    unmatched_headers: list[str] = field(default_factory=list)


def _detect_header_row(ws, max_scan: int = 5) -> int:
    """Return the 1-indexed row number most likely to be the header row."""
    best_row, best_score = 1, -1
    for r in range(1, min(max_scan, ws.max_row) + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        non_empty = [c for c in cells if c not in (None, "")]
        if not non_empty:
            continue
        text_like = sum(1 for c in non_empty if isinstance(c, str))
        score = text_like / len(non_empty) * len(non_empty)
        if score > best_score:
            best_score, best_row = score, r
    return best_row


def _first_empty_row(ws, header_row: int) -> int:
    """First row (after header_row) where every cell is blank."""
    r = header_row + 1
    while r <= ws.max_row:
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v in (None, "") for v in row_vals):
            return r
        r += 1
    return max(r, header_row + 1)


def _tokenize(text: str) -> list[str]:
    """Split a header/key into lowercase word tokens, handling camelCase,
    snake_case, kebab-case, and spaces. 'panNumber' -> ['pan', 'number']."""
    s = str(text)
    s = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", s)  # camelCase boundary
    s = re.sub(r"[^A-Za-z0-9]+", " ", s)
    return [t.lower() for t in s.split() if t]


# ---------------------------------------------------------------------------
# Name-field handling (Full Name <-> First Name / Last Name)
# ---------------------------------------------------------------------------
# NOTE: "first", "last", "middle", "given" live in QUALIFIER_WORDS so that
# generic fuzzy matching still recognises e.g. "First Name" as basically a
# "name" field. But that same stripping meant a header's "part" (first/
# last/full) was invisible to the qualifier-stripped-match check, so
# "Full Name", "First Name" and "Last Name" could all collapse onto
# whichever single name-like key happened to be matched first - stealing
# each other's data or duplicating one value into every name column. The
# helpers below explicitly detect which "part" a header/key represents and
# synthesize the missing counterpart (a combined full name from separate
# first/last keys, or a split first/last from a combined full name key)
# BEFORE fuzzy matching runs, so each column gets the right value via a
# same-name exact match instead of guesswork.
_FIRST_NAME_SYNONYMS = {"first", "given", "fname", "forename"}
_LAST_NAME_SYNONYMS = {"last", "surname", "family", "lname"}
_MIDDLE_NAME_SYNONYMS = {"middle", "mname"}
_FULL_NAME_QUALIFIERS = (
    QUALIFIER_WORDS - _FIRST_NAME_SYNONYMS - _LAST_NAME_SYNONYMS - _MIDDLE_NAME_SYNONYMS
)

# Honorifics/titles that precede a name and suffixes that follow it - neither
# is part of the person's actual first/last name, so both are peeled off
# before splitting (a naive split would otherwise turn "Dr. Ramesh Kumar
# Sharma" into first="Dr." last="Sharma", which is wrong).
_NAME_PREFIXES = {
    "mr", "mrs", "ms", "miss", "mx", "dr", "prof", "er",
    "shri", "smt", "kumari", "master", "sri", "capt", "col", "gen", "rev",
}
_NAME_SUFFIXES = {
    "jr", "sr", "ii", "iii", "iv", "v", "phd", "md", "esq",
}


def _name_role(tokens: frozenset) -> str | None:
    """Classify a header/key's word tokens as a name 'part', if any."""
    if tokens & _FIRST_NAME_SYNONYMS:
        return "first"
    if tokens & _LAST_NAME_SYNONYMS:
        return "last"
    if tokens & _MIDDLE_NAME_SYNONYMS:
        return "middle"
    if "name" in tokens:
        remaining = tokens - _FULL_NAME_QUALIFIERS
        if remaining == {"name"}:
            return "full"
    return None


def _split_full_name(value: str) -> tuple[str, str, str]:
    """
    Split a real-world full name into (first, middle, last).

    Handles the cases that actually show up in real data, not just a plain
    "First Last":
    - Titles/honorifics up front ("Dr. Ramesh Kumar Sharma" -> title dropped,
      not mistaken for the first name).
    - Generational/professional suffixes at the end ("John Doe Jr.",
      "Rakesh Verma III" -> kept attached to the last name, not treated as
      a separate middle word).
    - Multiple middle names ("Mary Jane Watson Parker" -> first="Mary",
      middle="Jane Watson", last="Parker").
    - A single-word name ("Madonna" -> first="Madonna", no last name -
      nothing invented).
    - Extra whitespace/commas from messy source data.
    """
    words = [w.strip(" ,") for w in value.strip().split() if w.strip(" ,")]
    if not words:
        return "", "", ""

    if len(words) > 1 and words[0].rstrip(".").lower() in _NAME_PREFIXES:
        words = words[1:]
    if not words:
        return "", "", ""

    suffix = ""
    if len(words) > 1 and words[-1].rstrip(".").lower() in _NAME_SUFFIXES:
        suffix = words[-1]
        words = words[:-1]
    if not words:
        return "", "", suffix

    first = words[0]
    if len(words) == 1:
        return first, "", suffix
    last = words[-1]
    if suffix:
        last = f"{last} {suffix}"
    middle = " ".join(words[1:-1])
    return first, middle, last


def _augment_name_fields(record: dict[str, Any]) -> dict[str, Any]:
    """
    Detect first/last/middle/full name keys already present in a record -
    however they're actually spelled (given_name, SURNAME, customerName,
    ...) - and:
    1. Alias each one to its canonical name (firstName/middleName/
       lastName/fullName) so a plainly-worded Excel header like "First
       Name" can match it directly, instead of only matching keys that
       happen to already be spelled that way.
    2. Synthesize whichever combined-or-split counterpart is genuinely
       missing: first+last (no full) -> a joined "fullName"; full (no
       first/last) -> "firstName"/"middleName"/"lastName" split via
       _split_full_name.
    Never overwrites a value that's already present under the canonical
    key - real data always wins over anything synthesized.
    """
    roles: dict[str, str] = {}
    for key in record:
        leaf = key.split(".")[-1]
        leaf = re.sub(r"\[\d+\]$", "", leaf)
        role = _name_role(frozenset(_tokenize(leaf)))
        if role and role not in roles:
            roles[role] = key

    if not roles:
        return record

    augmented = dict(record)
    canonical_key = {"first": "firstName", "middle": "middleName", "last": "lastName", "full": "fullName"}
    for role, key in roles.items():
        canon = canonical_key[role]
        value = record.get(key)
        if canon not in augmented and value not in (None, ""):
            augmented[canon] = value

    first_k, last_k = roles.get("first"), roles.get("last")
    middle_k, full_k = roles.get("middle"), roles.get("full")

    if (first_k or last_k) and not full_k:
        parts = [
            str(record[k]).strip()
            for k in (first_k, middle_k, last_k)
            if k and record.get(k) not in (None, "") and str(record[k]).strip()
        ]
        if parts:
            augmented["fullName"] = " ".join(parts)

    if full_k and not (first_k and last_k):
        first, middle, last = _split_full_name(str(record.get(full_k, "")))
        if not first_k and first:
            augmented["firstName"] = first
        if not middle_k and middle:
            augmented["middleName"] = middle
        if not last_k and last:
            augmented["lastName"] = last

    return augmented


def _best_match(header: str, record: dict[str, Any]) -> str | None:
    """
    Find the record key that best matches an Excel header. Returns None if
    nothing clears the confidence bar.

    Strategy (in order of trust):
    1. Exact match on the normalized full string.
    2. Acronym match - e.g. header "dob" matches key "date_of_birth" because
       its word-initials spell "dob".
    3. Token overlap - header and key are split into words (camelCase/
       snake_case/space aware) and matched by how much of the *smaller*
       token set is covered by shared words. This is what lets "Email
       Address" match "email" and "Customer Name" match "customerName",
       while correctly rejecting things like "panNumber" vs "mobileNumber"
       (they only share the word "number", not enough coverage).
    4. Plain character-similarity fallback - only for single-word vs
       single-word comparisons, to tolerate small typos/spelling variants.
    """
    norm_header = normalize(header)
    if not norm_header:
        return None

    header_tokens = _tokenize(header)
    best_key, best_score = None, 0

    for key in record:
        leaf = key.split(".")[-1]
        leaf = re.sub(r"\[\d+\]$", "", leaf)
        norm_leaf = normalize(leaf)
        norm_full = normalize(key)

        if norm_header == norm_leaf or norm_header == norm_full:
            return key  # exact match, short circuit

        leaf_tokens = _tokenize(leaf)

        # Acronym match: "dob" <-> ["date", "of", "birth"]
        if len(leaf_tokens) >= 2 and norm_header == "".join(t[0] for t in leaf_tokens):
            return key
        if len(header_tokens) >= 2 and norm_leaf == "".join(t[0] for t in header_tokens):
            return key

        # Qualifier-stripped match: drop generic qualifier words ("full",
        # "customer", "contact", ...) from BOTH sides and check what's left
        # is exactly equal. This is what lets "Customer Name" match
        # "personalDetails.fullName" (both reduce to just {"name"}), while
        # still correctly rejecting "Product Name" (reduces to
        # {"product", "name"} - "product" isn't a qualifier, so it stays
        # and breaks the equality, leaving it unmatched as it should be).
        h_core = frozenset(t for t in header_tokens if t not in QUALIFIER_WORDS)
        l_core = frozenset(t for t in leaf_tokens if t not in QUALIFIER_WORDS)
        if h_core and h_core == l_core:
            return key

        # Token overlap: how much of the smaller word-set is shared
        hset, lset = set(header_tokens), set(leaf_tokens)
        overlap = hset & lset
        token_score = 0
        if overlap:
            coverage = len(overlap) / min(len(hset), len(lset))
            if coverage >= 0.6:
                token_score = coverage * 100

        # Character-similarity fallback: only trusted when BOTH sides are a
        # single word each, so we don't get fooled by shared word-fragments
        # across multi-word keys (e.g. "...Number" suffix collisions).
        # Uses whole-string ratio only (NOT partial_ratio) - partial_ratio
        # scores on the best-aligned substring, which lets short, wholly
        # unrelated words score deceptively high off a shared fragment
        # (e.g. "quantity" vs "city" scores 86% on partial_ratio thanks to
        # the shared "ity" tail, but only 50% on the full-word ratio, which
        # correctly reflects that they aren't the same word).
        str_score = 0
        if len(header_tokens) == 1 and len(leaf_tokens) == 1 and min(len(norm_header), len(norm_leaf)) >= 4:
            str_score = fuzz.ratio(norm_header, norm_leaf)

        score = max(token_score, str_score)
        if score > best_score:
            best_score, best_key = score, key

    if best_score >= FUZZY_THRESHOLD:
        return best_key
    return None


# Header tokens that usually mean "this column is a unique identifier" -
# checked in priority order (most trustworthy first) when deciding which
# column(s) to use for spotting a duplicate record.
_IDENTIFIER_TOKENS = [
    "uid", "mrid", "mrn", "aadhar", "aadhaar", "pan", "id",
    "email", "account", "number", "code", "mobile", "phone",
]


def _pick_key_columns(headers: dict[int, str], matches: dict[int, str]) -> list[int]:
    """
    Decide which matched column(s) identify "the same record" for duplicate
    detection. Prefers a single obviously-unique column (uid, aadhaar, pan,
    id, email, account/number/code, mobile/phone - in that trust order).
    Falls back to using ALL matched columns together (composite key) when no
    such column is present.
    """
    for token in _IDENTIFIER_TOKENS:
        for col in matches:
            if token in _tokenize(headers.get(col, "")):
                return [col]
    return list(matches.keys())


def _norm_cell(value: Any) -> str:
    """Normalize a cell value for duplicate comparison."""
    if value in (None, ""):
        return ""
    return str(value).strip().lower()


def _existing_row_index(ws, header_row: int, key_cols: list[int]) -> dict[tuple, int]:
    """
    Scan already-filled rows below the header and build a lookup of
    key-values -> row number, so a new record that matches an existing row's
    key can REPLACE it instead of being appended as a fresh duplicate row.
    """
    index: dict[tuple, int] = {}
    if not key_cols:
        return index

    r = header_row + 1
    while r <= ws.max_row:
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v not in (None, "") for v in row_vals):
            key = tuple(_norm_cell(ws.cell(row=r, column=c).value) for c in key_cols)
            if any(key):  # ignore all-blank/N-A keys, nothing reliable to match on
                index[key] = r
        r += 1
    return index


def build_plan(ws, record: dict[str, Any]) -> SheetPlan:
    """Detect headers for one sheet and map them against one record."""
    header_row = _detect_header_row(ws)
    plan = SheetPlan(sheet_name=ws.title, header_row=header_row)

    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val in (None, ""):
            continue
        header_text = str(val)
        plan.headers[c] = header_text
        match_key = _best_match(header_text, record)
        if match_key is not None:
            plan.matches[c] = match_key
        else:
            plan.unmatched_headers.append(header_text)

    return plan


def fill_workbook(
    excel_path: str,
    data_path: str,
    save_path: str | None = None,
    skip_unrelated_sheets: bool = True,
) -> dict:
    """
    Main entry point.

    - A response only fills the sheet(s) it actually relates to. If a
      workbook has many sheets (e.g. an e-commerce file with "Customers",
      "Orders", "Products"...), a customer-details response fills only the
      matching sheet - other sheets stay untouched. Feed a different
      response later (Orders API, Products API, ...) and IT fills its own
      matching sheet, same workbook. Set `skip_unrelated_sheets=False` to
      instead force every sheet to get a row (unmatched cells become "N/A").
    - Any header with no corresponding value in the response is written as
      "N/A".
    - Duplicate records REPLACE the existing row instead of adding a new
      one. "Duplicate" is decided by a unique-looking column if the sheet
      has one (uid / aadhaar / pan / id / email / account / number / code /
      mobile - whichever appears first), otherwise by all matched columns
      together. A genuinely new record still gets a brand new row.

    Returns a summary dict describing what was matched/filled/replaced per
    sheet, for display in the UI / logs. Saves the workbook to `save_path`
    (defaults to overwriting `excel_path`).
    """
    save_path = save_path or excel_path
    records = parse_data_file(data_path)
    if not records:
        raise ValueError("No records found in the JSON/XML file.")
    records = [_augment_name_fields(r) for r in records]

    wb = openpyxl.load_workbook(excel_path)
    summary = {"records_processed": len(records), "sheets": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row == 0 or ws.max_column == 0:
            continue

        sheet_summary = {
            "sheet": sheet_name,
            "matched_headers": {},
            "unmatched_headers": [],
            "rows_filled": 0,
            "rows_added": 0,
            "rows_replaced": 0,
            "skipped": False,
        }

        # Use the first record just to detect the header row / column layout
        # and to judge whether this sheet is even relevant to this response.
        first_plan = build_plan(ws, records[0])
        if not first_plan.headers:
            # Truly empty header row - nothing to key off, leave it alone.
            sheet_summary["skipped"] = True
            summary["sheets"].append(sheet_summary)
            continue

        sheet_is_relevant = any(build_plan(ws, r).matches for r in records)
        if skip_unrelated_sheets and not sheet_is_relevant:
            sheet_summary["skipped"] = True
            sheet_summary["unmatched_headers"] = first_plan.unmatched_headers
            summary["sheets"].append(sheet_summary)
            continue

        header_row = first_plan.header_row
        key_cols = _pick_key_columns(first_plan.headers, first_plan.matches)
        key_to_row = _existing_row_index(ws, header_row, key_cols)
        next_row = _first_empty_row(ws, header_row)
        ever_matched: dict[str, str] = {}  # header_text -> key (union across all records)

        for record in records:
            plan = build_plan(ws, record)

            row_values: dict[int, Any] = {}
            for col_idx, header_text in plan.headers.items():
                if col_idx in plan.matches:
                    key = plan.matches[col_idx]
                    value = record.get(key)
                    if value is None or (isinstance(value, str) and value.strip() == ""):
                        value = BLANK_PLACEHOLDER
                    row_values[col_idx] = value
                    ever_matched[header_text] = key
                else:
                    row_values[col_idx] = BLANK_PLACEHOLDER

            record_key = tuple(_norm_cell(row_values.get(c)) for c in key_cols) if key_cols else ()
            target_row = key_to_row.get(record_key) if record_key and any(record_key) else None

            if target_row is not None:
                row = target_row
                sheet_summary["rows_replaced"] += 1
            else:
                row = next_row
                next_row += 1
                sheet_summary["rows_added"] += 1

            for col_idx, value in row_values.items():
                ws.cell(row=row, column=col_idx, value=value)

            if record_key and any(record_key):
                key_to_row[record_key] = row

            sheet_summary["rows_filled"] += 1

        sheet_summary["matched_headers"] = ever_matched
        sheet_summary["unmatched_headers"] = [
            h for h in first_plan.headers.values() if h not in ever_matched
        ]
        summary["sheets"].append(sheet_summary)

    try:
        wb.save(save_path)
        summary["saved_to"] = save_path
    except PermissionError:
        # Usually means the Excel file is currently open in Excel (or is
        # read-only). Don't lose the work - save alongside it instead.
        base, ext = os.path.splitext(save_path)
        fallback_path = f"{base}_filled{ext}"
        n = 2
        while os.path.exists(fallback_path):
            fallback_path = f"{base}_filled{n}{ext}"
            n += 1
        try:
            wb.save(fallback_path)
        except PermissionError as e:
            raise PermissionError(
                f"Could not save to '{save_path}' or '{fallback_path}'. "
                f"Close the Excel file if it's open, or check that the "
                f"Downloads folder allows write access, then try again."
            ) from e
        summary["saved_to"] = fallback_path
        summary["save_fallback"] = True

    return summary